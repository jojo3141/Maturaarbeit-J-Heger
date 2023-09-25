from class_Cubie import *
import statistics
default_tps = 7.5

# access the data

with open("solves_data/LBL_solutions.txt", "r") as file:
    all_LBL_solutions = [eval(line.strip()) for line in file]

with open("solves_data/CFOP_solutions.txt", "r") as file:
    all_CFOP_solutions = [eval(line.strip()) for line in file]

with open("solves_data/CFCE_solutions.txt", "r") as file:
    all_CFCE_solutions = [eval(line.strip()) for line in file]

with open("solves_data/ROUX_solutions.txt", "r") as file:
    all_ROUX_solutions = [eval(line.strip()) for line in file]

import numpy as np
import matplotlib.pyplot as plt


def create_diagram(data1, data2, data3, data4, xlabel="", ylabel="", title="",
                   save_in_place="solves_data/I_forgot_to_save_it", average=False):
    # Calculate the histogram values and bin edges
    hist_LBL, bins_LBL = np.histogram(data1, bins=200, density=False)
    hist_CFOP, bins_CFOP = np.histogram(data2, bins=200, density=False)
    hist_CFCE, bins_CFCE = np.histogram(data3, bins=200, density=False)
    hist_ROUX, bins_ROUX = np.histogram(data4, bins=200, density=False)

    # Calculate the total number of data points for normalization
    total_points = len(data1) + len(data2) + len(data3) + len(data4)

    # Normalize the histogram values
    hist_LBL = hist_LBL / total_points
    hist_CFOP = hist_CFOP / total_points
    hist_CFCE = hist_CFCE / total_points
    hist_ROUX = hist_ROUX / total_points

    # Create the area plots for each dataset
    plt.fill_between(bins_LBL[:-1], hist_LBL, color='orange', alpha=0.9, label='LBL', step='mid')
    plt.fill_between(bins_CFOP[:-1], hist_CFOP, color='green', alpha=0.9, label='CFOP', step='mid')
    plt.fill_between(bins_CFCE[:-1], hist_CFCE, color='yellow', alpha=0.9, label='CFCE', step='mid')
    plt.fill_between(bins_ROUX[:-1], hist_ROUX, color='pink', alpha=0.9, label='Roux', step='mid')

    # Add labels and legend
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    plt.title(title)
    plt.legend()

    # Add average information
    if average:
        avg_data1 = np.mean(data1)
        avg_data2 = np.mean(data2)
        avg_data3 = np.mean(data3)
        avg_data4 = np.mean(data4)

        print(avg_data1, avg_data2, avg_data3, avg_data4)

    # Save the plot
    plt.savefig(save_in_place, dpi=600)  # Set DPI to 600
    # Show the plot
    plt.show()


def move_count(solution):
    count = 0
    for elem in solution:
        if any(char.isupper() for char in elem):
            count += 1
            if "2" in elem:
                count += 1
        elif elem in ["x", "xi", "y", "yi"]:
            count += 1
    return count# + random.random()/2
def create_move_count_diagram():
    all_LBL_move_counts  = [move_count(solution) for solution in all_LBL_solutions]
    all_CFOP_move_counts = [move_count(solution) for solution in all_CFOP_solutions]
    all_CFCE_move_counts = [move_count(solution) for solution in all_CFCE_solutions]
    all_ROUX_move_counts = [move_count(solution) for solution in all_ROUX_solutions]
    create_diagram(all_LBL_move_counts, all_CFOP_move_counts, all_CFCE_move_counts, all_ROUX_move_counts, "Zuganzahl pro Lösung", "Häufigkeit", "Wahrscheinlichkeit der Zuganzahl pro Lösung", "solves_data/move_count_histogramm2.png", average=True)


def estimate_time_for_solution(solution, tps):
    relative_R = 1
    relative_R2 = 1.5
    relative_L = 1.1
    relative_L2 = 1.6
    relative_F = 1.2
    relative_F2 = 1.6
    relative_B = 1.3
    relative_B2 = 2.2
    relative_x = 2
    relative_M = 2.7
    relative_Mi = 3.2
    relative_M2 = 3.9
    relative_M_in_MUseries = 2.4
    relative_Mi_in_MUseries = 2.9
    relative_M2_in_MUseries = 3.5
    relative_R_in_RUseries = 0.7
    relative_special_double_turns = 1.7
    relative_difficult_look_ahead = 2
    relative_easy_look_ahead = 0.8

    time = 0
    for i in range(len(solution)):
        turn = solution[i]

        # RUturns in series
        if turn in ["R", "Ri", "U", "Ui"] and\
                (all(elem in ["U", "Ui", "U2", "R", "Ri", "R2"] for elem in solution[i - 2:i])):
            time += 1 / tps * relative_R_in_RUseries

        # Special double turns
        elif solution[i - 1:i] in [["U", "D"], ["Ui", "D"], ["U", "Di"], ["Ui", "Di"],
                                   ["M", "Ri"], ["Ri", "M"], ["Mi", "R"], ["R", "Mi"],
                                   ["M", "Li"], ["Li", "M"], ["Mi", "L"], ["L", "Mi"]]:
            time += (1 / tps) * (relative_special_double_turns / 2)

        # Normal single turns independent
        elif turn in ["R", "Ri", "U", "Ui"]:
            time += 1 / tps * relative_R
        elif turn in ["L", "Li", "D", "Di"]:
            time += 1 / tps * relative_L
        elif turn in ["F", "Fi"]:
            time += 1 / tps * relative_F
        elif turn in ["B", "Bi"]:
            time += 1 / tps * relative_B
        elif "y" in turn or "x" in turn:
            time += 1 / tps * relative_x

        # Normal doulbe turns independent
        elif turn in ["L2", "D2"]:
            time += 1 / tps * relative_L2
        elif turn in ["R2", "U2"]:
            time += 1 / tps * relative_R2
        elif turn == "F2":
            time += 1 / tps * relative_F2
        elif turn == "B2":
            time += 1 / tps * relative_B2

        # Mturns in series
        elif turn == "M" and (all(elem in ["U", "Ui", "U2", "M", "Mi", "M2"] for elem in solution[i - 2:i])):
            time += 1 / tps * relative_M_in_MUseries
        elif turn == "Mi" and (all(elem in ["U", "Ui", "U2", "M", "Mi", "M2"] for elem in solution[i - 2:i])):
            time += 1 / tps * relative_Mi_in_MUseries
        elif turn == "M2" and (all(elem in ["U", "Ui", "U2", "M", "Mi", "M2"] for elem in solution[i - 2:i])):
            time += 1 / tps * relative_M2_in_MUseries

        # Mturns independent
        elif turn == "M":
            time += 1 / tps * relative_M
        elif turn == "Mi":
            time += 1 / tps * relative_Mi
        elif turn == "M2":
            time += 1 / tps * relative_M2

        # Difficult Look Ahead
        elif turn == " | ":
            if any(elem in ["M", "Mi", "M2", "B", "Bi", "B2"] for elem in solution[i-5:i]):
                time += 3 / tps * relative_difficult_look_ahead
            else:
                time += 1 / tps * relative_easy_look_ahead
    return time

def create_time_diagram():
    all_LBL_times  = [estimate_time_for_solution(solution, tps=default_tps) for solution in all_LBL_solutions]
    all_CFOP_times = [estimate_time_for_solution(solution, tps=default_tps) for solution in all_CFOP_solutions]
    all_CFCE_times = [estimate_time_for_solution(solution, tps=default_tps) for solution in all_CFCE_solutions]
    all_ROUX_times = [estimate_time_for_solution(solution, tps=default_tps) for solution in all_ROUX_solutions]
    print(sum(all_LBL_times)/len(all_LBL_times))
    print(sum(all_CFOP_times) / len(all_CFOP_times))
    print(sum(all_CFCE_times) / len(all_CFCE_times))
    print(sum(all_ROUX_times) / len(all_ROUX_times))
    stddev_LBL = statistics.stdev(all_LBL_times)
    stddev_CFOP = statistics.stdev(all_CFOP_times)
    stddev_CFCE = statistics.stdev(all_CFCE_times)
    stddev_ROUX = statistics.stdev(all_ROUX_times)

    print(f"Standard Deviation LBL Time: {stddev_LBL}")
    print(f"Standard Deviation CFOP Time: {stddev_CFOP}")
    print(f"Standard Deviation CFCE Time: {stddev_CFCE}")
    print(f"Standard Deviation ROUX Time: {stddev_ROUX}")
    create_diagram(all_LBL_times, all_CFOP_times, all_CFCE_times, all_ROUX_times, "Lösungszeit in Sekunden", "Relative Häufigkeitsdichte", "Verteilung der Lösungszeit", "solves_data/time_histogram10.png")

def time_per_step(solution):
    step_list, turn_list, time_list = [], [], []
    for i in range(len(solution)):
        if solution[i] == " | ":
            step_list.append(i)
    for i in range(len(step_list)):
        if i != len(step_list)-1:
            turn_list.append(solution[step_list[i]+1:step_list[i+1]])
        else:
            turn_list.append(solution[step_list[i]+1:])
    #print(turn_list)
    for elem in turn_list:
        time_list.append(estimate_time_for_solution(elem, default_tps))
    return time_list

def generate_data_for_step_diagram():
    sum_list = [0, 0, 0, 0, 0, 0, 0, 0]
    for solution in all_CFOP_solutions:
        step_list = time_per_step(solution)
        if len(step_list) == 7:
            new_list = []
            n = len(solution)
            for i in range(n):
                new_list.append(solution[i])
                if i < n - 1 and solution[i] == solution[i + 1]:
                    new_list.append(solution[i])
            solution = new_list
            step_list = time_per_step(solution)
        for i in range(8):
            sum_list[i] += step_list[i]

    sum_list = [elem/10000 for elem in sum_list]
    pc = [(elem / sum(sum_list)) * 100 for elem in sum_list]
    percent_list = [pc[0], pc[1] + pc[2] + pc[3] + pc[4], pc[6], pc[7]]
    print("CFOP:", percent_list)

    sum_list = [0, 0, 0, 0, 0, 0, 0, 0]
    for solution in all_CFCE_solutions:
        step_list = time_per_step(solution)
        if len(step_list) == 7:
            new_list = []
            n = len(solution)
            for i in range(n):
                new_list.append(solution[i])
                if i < n - 1 and solution[i] == solution[i + 1]:
                    new_list.append(solution[i])
            solution = new_list
            step_list = time_per_step(solution)
        for i in range(8):
            sum_list[i] += step_list[i]

    sum_list = [elem/10000 for elem in sum_list]
    pc = [(elem / sum(sum_list)) * 100 for elem in sum_list]
    percent_list = [pc[0], pc[1] + pc[2] + pc[3] + pc[4], pc[6], pc[7]]
    print("CFCE:", percent_list)

    sum_list = [0, 0, 0, 0, 0, 0, 0]
    for solution in all_LBL_solutions:
        step_list = time_per_step(solution)
        for i in range(7):
            sum_list[i] += step_list[i]

    sum_list = [elem / 10000 for elem in sum_list]
    percent_list = [(elem / sum(sum_list)) * 100 for elem in sum_list]
    print("LBL:", percent_list)

    sum_list = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    for solution in all_ROUX_solutions:
        step_list = time_per_step(solution)
        if len(step_list) == 9:
            new_list = []
            n = len(solution)
            for i in range(n):
                new_list.append(solution[i])
                if i < n - 1 and solution[i] == solution[i + 1]:
                    new_list.append(solution[i])
            solution = new_list
            step_list = time_per_step(solution)
        for i in range(10):
            sum_list[i] += step_list[i]

    sum_list = [elem / 10000 for elem in sum_list]
    pc = [(elem / sum(sum_list)) * 100 for elem in sum_list]
    percent_list = [pc[0] + pc[1] + pc[2] + pc[3] + pc[4], pc[6], pc[7],  pc[8] + pc[9]]
    print("ROUX:", percent_list)


def count_rotations_and_Mturns(solution):
    Mcount, Rcount, UR_count, FLBD_count = 0, 0, 0, 0
    for elem in solution:
        if elem in ["M", "Mi"]:
            Mcount += 1
        elif elem == "M2":
            Mcount += 2
        elif elem in ["x", "xi", "y", "yi"]:
            Rcount += 1
        elif elem in ["U", "Ui", "R", "Ri"]:
            UR_count += 1
        elif elem in ["U2", "R2"]:
            UR_count += 2
        elif elem in normal_moves + inverted_moves:
            FLBD_count += 1
        elif elem in double_moves:
            FLBD_count += 2

    return [Mcount, Rcount, UR_count, FLBD_count]

def generate_data_for_rotations_diagram():
    sum_list = [0, 0, 0, 0]
    for solution in all_LBL_solutions:
        m_and_r_data = count_rotations_and_Mturns(solution)
        for i in range(4):
            sum_list[i] += m_and_r_data[i]
    sum_list = [elem/10000 for elem in sum_list]
    print("LBL:", sum_list, sum(sum_list))

    sum_list = [0, 0, 0, 0]
    for solution in all_CFOP_solutions:
        m_and_r_data = count_rotations_and_Mturns(solution)
        for i in range(4):
            sum_list[i] += m_and_r_data[i]
    sum_list = [elem / 10000 for elem in sum_list]
    print("CFOP:", sum_list, sum(sum_list))

    sum_list = [0, 0, 0, 0]
    for solution in all_CFCE_solutions:
        m_and_r_data = count_rotations_and_Mturns(solution)
        for i in range(4):
            sum_list[i] += m_and_r_data[i]
    sum_list = [elem / 10000 for elem in sum_list]
    print("CFCE:", sum_list, sum(sum_list))

    sum_list = [0, 0, 0, 0]
    for solution in all_ROUX_solutions:
        m_and_r_data = count_rotations_and_Mturns(solution)
        for i in range(4):
            sum_list[i] += m_and_r_data[i]
    sum_list = [elem / 10000 for elem in sum_list]
    print("ROUX:", sum_list, sum(sum_list))



#create_move_count_diagram()

create_time_diagram()
#generate_data_for_step_diagram()

#generate_data_for_rotations_diagram()

# how many steps
# how easy is the look-ahead

# time per step?
# Inspection time?
# Difficulty of the algorithms



"""
import openpyxl

# Erstelle eine neue Excel-Datei
workbook = openpyxl.Workbook()

# Erstelle ein neues Arbeitsblatt
sheet = workbook.active

# Vier Listen mit Zahlen


liste1 = [move_count(solution) for solution in all_LBL_solutions]
liste2 = [move_count(solution) for solution in all_CFOP_solutions]
liste3 = [move_count(solution) for solution in all_CFCE_solutions]
liste4 = [move_count(solution) for solution in all_ROUX_solutions]

def count_occurrences(input_list):
    # Create a list of length 251 (0-250) initialized with zeros
    result_list = [0] * 251

    # Count the occurrences of each number in the input list
    for num in input_list:
        if 0 <= num <= 250:
            result_list[num] += 1

    return result_list


def insert_zeros_after_even(input_list):
    result_list = []
    for i, num in enumerate(input_list):
        result_list.append(num)
        result_list.append(0)
    return result_list


# Example usage:
input_list = [1, 2, 3, 4, 5]
output_list = insert_zeros_after_even(input_list)
print(output_list)

liste1 = insert_zeros_after_even(count_occurrences(liste1))
liste2 = insert_zeros_after_even(count_occurrences(liste2))
liste3 = insert_zeros_after_even(count_occurrences(liste3))
liste4 = insert_zeros_after_even(count_occurrences(liste4))

# Example usage:
input_list = [1, 2, 3, 2, 1, 4, 5, 5, 5]
output_list = count_occurrences(input_list)
print(output_list)


# Schreibe die Zahlen in das Arbeitsblatt
for row, (val1, val2, val3, val4) in enumerate(zip(liste1, liste2, liste3, liste4), start=2):
    sheet.cell(row=row, column=2, value=val1)
    sheet.cell(row=row, column=3, value=val2)
    sheet.cell(row=row, column=4, value=val3)
    sheet.cell(row=row, column=5, value=val4)

# Speichere die Excel-Datei
workbook.save("Time_of_10000_solves4.xlsx")

# Schließe die Arbeitsmappe
workbook.close()

print("Excel-Datei wurde erstellt und Zahlen wurden eingetragen.")

"""